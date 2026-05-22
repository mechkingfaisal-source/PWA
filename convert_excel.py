"""
convert_excel.py  (v2 — with Engineer + Technician support)
════════════════════════════════════════════════════════════════
Reads your Excel file (log_in.xlsx) which should have TWO sheets:

    Sheet 1 → "Technicians"   columns: Iqama Number | File Number | Name
    Sheet 2 → "Engineers"     columns: Iqama Number | File Number | Name

Then it generates the JavaScript arrays to paste into login.html.

USAGE:
    python convert_excel.py

REQUIREMENTS:
    pip install openpyxl

HOW TO USE THE OUTPUT:
    1. Run this script
    2. It shows two blocks: one for TECHNICIANS, one for ENGINEERS
    3. Open login.html
    4. Find TECHNICIANS = [ ... ] → paste Technician block inside
    5. Find ENGINEERS   = [ ... ] → paste Engineer block inside
    6. Save and done!

════════════════════════════════════════════════════════════════
  EXCEL SHEET SETUP GUIDE
════════════════════════════════════════════════════════════════

Your Excel file (log_in.xlsx) needs 2 sheets like this:

  ┌─────────────────────────────────────────────────────────┐
  │  Sheet name: Technicians                                │
  │  ─────────────────────────────────────────────────────  │
  │  Iqama Number  │  File Number  │  Name                  │
  │  2123456789    │  F001         │  Ahmed Al-Rashidi      │
  │  2987654321    │  F002         │  Mohammed Al-Zahrani   │
  └─────────────────────────────────────────────────────────┘

  ┌─────────────────────────────────────────────────────────┐
  │  Sheet name: Engineers                                  │
  │  ─────────────────────────────────────────────────────  │
  │  Iqama Number  │  File Number  │  Name                  │
  │  1023456789    │  E001         │  Eng. Faisal Al-Otaibi │
  │  1087654321    │  E002         │  Eng. Omar Al-Shehri   │
  └─────────────────────────────────────────────────────────┘

════════════════════════════════════════════════════════════════
"""

import sys, os

try:
    import openpyxl
except ImportError:
    print("  Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

# ── CONFIGURATION ─────────────────────────────────────────────────────────────
EXCEL_FILE   = "log_in.xlsx"

# Sheet names in your Excel file:
TECH_SHEET   = "Technicians"
ENG_SHEET    = "Engineers"

# Column header names (must match what's in your Excel exactly):
IQAMA_COL    = "Iqama Number"
FILE_COL     = "File Number"
NAME_COL     = "Name"            # optional — leave blank in Excel if not needed
# ──────────────────────────────────────────────────────────────────────────────


def find_columns(headers, iqama_col, file_col, name_col):
    """Return dict of column_name → index, case-insensitive."""
    hmap = {str(h).strip().lower(): i for i, h in enumerate(headers) if h}
    iq_idx = next((i for k, i in hmap.items() if iqama_col.lower() in k), None)
    fi_idx = next((i for k, i in hmap.items() if file_col.lower()  in k), None)
    nm_idx = next((i for k, i in hmap.items() if name_col.lower()  in k), None)
    return iq_idx, fi_idx, nm_idx


def read_sheet(ws, label):
    """Read a sheet and return list of credential dicts."""
    records = []
    headers = None
    header_row = None

    for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if any(IQAMA_COL.lower() in v.lower() for v in vals) or \
           any(FILE_COL.lower()  in v.lower() for v in vals):
            headers   = vals
            header_row = i
            break

    if headers is None:
        print(f"\n  ⚠  Could not find headers in sheet '{label}'.")
        print(f"     Make sure row 1 has: '{IQAMA_COL}', '{FILE_COL}', '{NAME_COL}'")
        print(f"     Or update IQAMA_COL / FILE_COL / NAME_COL at the top of this script.\n")
        return []

    iq_idx, fi_idx, nm_idx = find_columns(headers, IQAMA_COL, FILE_COL, NAME_COL)

    if iq_idx is None or fi_idx is None:
        print(f"\n  ⚠  Columns not matched in sheet '{label}'.")
        print(f"     Found headers: {headers}")
        return []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        iq = row[iq_idx] if iq_idx < len(row) else None
        fi = row[fi_idx] if fi_idx < len(row) else None
        nm = row[nm_idx] if (nm_idx is not None and nm_idx < len(row)) else None

        if iq and fi:
            records.append({
                "iqama": str(iq).strip(),
                "file":  str(fi).strip(),
                "name":  str(nm).strip() if nm else "",
            })

    return records


def to_js(records):
    """Convert list of dicts to JS array entries."""
    lines = []
    for r in records:
        name_part = f', name: "{r["name"]}"' if r["name"] else ''
        lines.append(f'      {{ iqama: "{r["iqama"]}", file: "{r["file"]}"{name_part} }},')
    return "\n".join(lines)


def print_block(label, var_name, records):
    print(f"\n  {'═'*60}")
    print(f"  {label}  ({len(records)} record{'s' if len(records)!=1 else ''})")
    print(f"  {'─'*60}")
    print(f"  Paste this inside  {var_name} = [  ...  ]  in login.html:\n")
    if records:
        print(to_js(records))
    else:
        print("      // (no records found)")
    print(f"  {'═'*60}")


def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"\n  ERROR: '{EXCEL_FILE}' not found in this folder.")
        print("  Make sure the Excel file is in the same directory as this script.\n")
        sys.exit(1)

    print(f"\n  Reading '{EXCEL_FILE}'...")
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    # ── Check which sheets exist ──────────────────────────────────────────────
    available = wb.sheetnames
    print(f"  Sheets found: {available}")

    if TECH_SHEET not in available:
        print(f"\n  ⚠  Sheet '{TECH_SHEET}' not found.")
        print(f"     Available sheets: {available}")
        print(f"     Rename your technician sheet to '{TECH_SHEET}' in Excel,")
        print(f"     OR update TECH_SHEET in this script to match.\n")
        tech_records = []
    else:
        tech_records = read_sheet(wb[TECH_SHEET], TECH_SHEET)

    if ENG_SHEET not in available:
        print(f"\n  ⚠  Sheet '{ENG_SHEET}' not found.")
        print(f"     Available sheets: {available}")
        print(f"     Rename your engineer sheet to '{ENG_SHEET}' in Excel,")
        print(f"     OR update ENG_SHEET in this script to match.\n")
        eng_records = []
    else:
        eng_records = read_sheet(wb[ENG_SHEET], ENG_SHEET)

    # ── Print output ──────────────────────────────────────────────────────────
    print_block("TECHNICIAN DATA  →  TECHNICIANS array", "TECHNICIANS", tech_records)
    print_block("ENGINEER DATA    →  ENGINEERS array",   "ENGINEERS",   eng_records)

    # ── Save to file ──────────────────────────────────────────────────────────
    out = "credentials_output.txt"
    with open(out, "w", encoding="utf-8") as f:
        f.write("// ── TECHNICIANS ──────────────────────────────────────────────\n")
        f.write("// Paste inside TECHNICIANS = [ ... ] in login.html\n\n")
        f.write(to_js(tech_records) or "      // (no records)")
        f.write("\n\n")
        f.write("// ── ENGINEERS ────────────────────────────────────────────────\n")
        f.write("// Paste inside ENGINEERS = [ ... ] in login.html\n\n")
        f.write(to_js(eng_records) or "      // (no records)")
        f.write("\n")

    print(f"\n  ✔ Output also saved to '{out}'")
    print("  Open login.html and paste each block into the matching array.\n")


if __name__ == "__main__":
    main()
